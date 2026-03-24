"""Excel解析服务"""

import openpyxl
from openpyxl.styles import Font
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel文件解析器"""

    # 测试用例字段映射
    FIELD_MAPPING = {
        "用例名称": "name",
        "用例描述": "description",
        "用例优先级": "priority",
        "用例类型": "case_type",
        "前置条件": "precondition",
        "测试步骤": "steps",
        "预期结果": "expected_result",
        "实际结果": "actual_result",
        "状态": "status",
        "测试人员": "tester",
        "测试日期": "test_date",
        "标签": "tags"
    }

    # 优先级映射
    PRIORITY_MAPPING = {
        "P0": "P0",
        "P1": "P1",
        "P2": "P2",
        "P3": "P3",
        "高": "P0",
        "中": "P1",
        "低": "P2",
        "high": "P0",
        "medium": "P1",
        "low": "P2"
    }

    # 用例类型映射
    CASE_TYPE_MAPPING = {
        "正向": "正向",
        "反向": "负向",
        "负向": "负向",
        "边界": "边界",
        "性能": "性能",
        "安全": "安全",
        "positive": "正向",
        "negative": "负向",
        "reverse": "负向",
        "edge": "边界",
        "performance": "性能",
        "security": "安全"
    }

    # 状态映射
    STATUS_MAPPING = {
        "待测试": "待测试",
        "测试中": "测试中",
        "通过": "通过",
        "失败": "失败",
        "阻塞": "阻塞",
        "跳过": "跳过",
        "pending": "待测试",
        "running": "测试中",
        "passed": "通过",
        "failed": "失败",
        "blocked": "阻塞",
        "skipped": "跳过"
    }

    @classmethod
    def parse_test_cases(cls, file_path: str) -> List[Dict[str, Any]]:
        """
        解析Excel文件中的测试用例

        Args:
            file_path: Excel文件路径

        Returns:
            测试用例列表
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            test_cases = []
            headers = []

            # 查找标题行（包含"用例名称"的行）
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for cell in row:
                    if cell and "用例名称" in str(cell):
                        header_row = row_idx
                        headers = [str(cell) if cell else "" for cell in row]
                        break
                if header_row:
                    break

            if not header_row:
                raise ValueError("未找到用例名称列，请检查Excel格式")

            # 构建列索引映射
            col_mapping = {}
            for col_idx, header in enumerate(headers):
                if header in cls.FIELD_MAPPING:
                    col_mapping[cls.FIELD_MAPPING[header]] = col_idx

            # 解析数据行
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                # 跳过空行
                if not row or not any(row):
                    continue

                test_case = cls._parse_test_case_row(row, col_mapping)
                if test_case and test_case.get("name"):
                    test_cases.append(test_case)

            workbook.close()
            logger.info(f"成功解析 {len(test_cases)} 条测试用例")
            return test_cases

        except Exception as e:
            logger.error(f"解析Excel文件失败: {str(e)}")
            raise

    @classmethod
    def _parse_test_case_row(cls, row: tuple, col_mapping: Dict[str, int]) -> Optional[Dict[str, Any]]:
        """解析单行测试用例数据"""
        test_case = {}

        # 必填字段：用例名称
        if "name" not in col_mapping:
            return None

        name_col = col_mapping["name"]
        if name_col >= len(row) or not row[name_col]:
            return None

        test_case["name"] = str(row[name_col]).strip()

        # 用例描述
        if "description" in col_mapping:
            col = col_mapping["description"]
            if col < len(row) and row[col]:
                test_case["description"] = str(row[col]).strip()

        # 优先级
        if "priority" in col_mapping:
            col = col_mapping["priority"]
            if col < len(row) and row[col]:
                priority_val = str(row[col]).strip()
                test_case["priority"] = cls.PRIORITY_MAPPING.get(priority_val, "P1")
            else:
                test_case["priority"] = "P1"
        else:
            test_case["priority"] = "P1"

        # 用例类型
        if "case_type" in col_mapping:
            col = col_mapping["case_type"]
            if col < len(row) and row[col]:
                type_val = str(row[col]).strip()
                test_case["case_type"] = cls.CASE_TYPE_MAPPING.get(type_val, "正向")
            else:
                test_case["case_type"] = "正向"
        else:
            test_case["case_type"] = "正向"

        # 前置条件
        if "precondition" in col_mapping:
            col = col_mapping["precondition"]
            if col < len(row) and row[col]:
                test_case["precondition"] = str(row[col]).strip()

        # 测试步骤（必填）
        if "steps" in col_mapping:
            col = col_mapping["steps"]
            if col < len(row) and row[col]:
                test_case["steps"] = str(row[col]).strip()
            else:
                test_case["steps"] = "1. 打开应用"
        else:
            test_case["steps"] = "1. 打开应用"

        # 预期结果（必填）
        if "expected_result" in col_mapping:
            col = col_mapping["expected_result"]
            if col < len(row) and row[col]:
                test_case["expected_result"] = str(row[col]).strip()
            else:
                test_case["expected_result"] = "操作成功"
        else:
            test_case["expected_result"] = "操作成功"

        # 实际结果
        if "actual_result" in col_mapping:
            col = col_mapping["actual_result"]
            if col < len(row) and row[col]:
                test_case["actual_result"] = str(row[col]).strip()

        # 状态
        if "status" in col_mapping:
            col = col_mapping["status"]
            if col < len(row) and row[col]:
                status_val = str(row[col]).strip()
                test_case["status"] = cls.STATUS_MAPPING.get(status_val, "待测试")
            else:
                test_case["status"] = "待测试"
        else:
            test_case["status"] = "待测试"

        # 测试人员
        if "tester" in col_mapping:
            col = col_mapping["tester"]
            if col < len(row) and row[col]:
                test_case["tester"] = str(row[col]).strip()

        # 测试日期
        if "test_date" in col_mapping:
            col = col_mapping["test_date"]
            if col < len(row) and row[col]:
                date_val = row[col]
                if isinstance(date_val, datetime):
                    test_case["test_date"] = date_val.strftime("%Y-%m-%d")
                else:
                    test_case["test_date"] = str(date_val).strip()

        # 标签
        if "tags" in col_mapping:
            col = col_mapping["tags"]
            if col < len(row) and row[col]:
                tags_str = str(row[col]).strip()
                test_case["tags"] = [tag.strip() for tag in tags_str.split(",") if tag.strip()]
            else:
                test_case["tags"] = []
        else:
            test_case["tags"] = []

        return test_case

    @classmethod
    def generate_excel_template(cls) -> bytes:
        """生成Excel导入模板"""
        from io import BytesIO

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "测试用例导入模板"

        # 设置标题行
        headers = [
            "用例名称", "用例描述", "用例优先级", "用例类型",
            "前置条件", "测试步骤", "预期结果", "实际结果",
            "状态", "测试人员", "测试日期", "标签"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # 添加示例数据
        example_data = [
            "登录功能-正常登录-001",
            "验证用户使用正确的用户名和密码可以成功登录",
            "P0",
            "正向",
            "1. 用户已注册\n2. 用户未登录",
            "1. 打开登录页面\n2. 输入正确的用户名和密码\n3. 点击登录按钮",
            "1. 登录成功\n2. 跳转到首页\n3. 显示用户信息",
            "",
            "待测试",
            "",
            "",
            "登录,基础功能"
        ]

        for col_idx, value in enumerate(example_data, 1):
            sheet.cell(row=2, column=col_idx, value=value)

        # 添加说明行
        sheet.cell(row=4, column=1, value="【字段说明】")
        sheet.cell(row=5, column=1, value="用例名称：必填，格式：模块-功能-编号")
        sheet.cell(row=6, column=1, value="用例优先级：可选值：P0(高)、P1(中)、P2(低)、P3")
        sheet.cell(row=7, column=1, value="用例类型：可选值：正向、负向、边界、性能、安全")
        sheet.cell(row=8, column=1, value="状态：可选值：待测试、测试中、通过、失败、阻塞、跳过")

        # 保存到字节流
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        workbook.close()

        return output.read()
